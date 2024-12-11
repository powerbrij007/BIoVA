import re #For caseignore
from re import IGNORECASE
from turtle import clear
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
import string

excel_file = pd.read_excel('RollList.xlsx')
excel_file[['Sno', 'Roll', 'Name']]
#print(excel_file)

#ex_at = pd.read_excel('Attendance.xlsx')
#ex_at[[]]
#print(ex_at)

wb = load_workbook('Attendance.xlsx')
#ws=wb.active #worksheet
#print(ws)
sheet = wb['20_Jan']

#getting sheet names
sheetName = wb.sheetnames
#print(sheetName[0])
for i in range(len(sheetName)):
    ws=wb[sheetName[i]] #sheet selected
    for row in ws.rows:
        for v in row:
            print(v.value)